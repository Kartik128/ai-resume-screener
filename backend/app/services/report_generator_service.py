import csv
import io
from typing import List, Sequence
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.models.job import Job
from app.models.score import Score


class ReportGeneratorService:
    """Service for generating CSV, Excel (XLSX), and PDF candidate reports."""

    @staticmethod
    def generate_csv_report(job: Job, scores: Sequence[Score]) -> str:
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            "Candidate Name",
            "Email",
            "Overall Score",
            "Mandatory Skills Score",
            "Experience Score",
            "Total Exp (Yrs)",
            "Location",
            "Date Evaluated",
        ])

        for s in scores:
            writer.writerow([
                s.candidate.full_name,
                s.candidate.email or "N/A",
                s.overall_score,
                s.mandatory_skills_score,
                s.experience_score,
                s.candidate.total_experience_years or 0.0,
                s.candidate.location or "N/A",
                s.created_at.strftime("%Y-%m-%d"),
            ])

        return output.getvalue()

    @staticmethod
    def generate_excel_report(job: Job, scores: Sequence[Score]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidate Leaderboard"

        # Styling
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        headers = [
            "Candidate Name",
            "Email",
            "Overall Score",
            "Mandatory Skills Score",
            "Experience Score",
            "Total Exp (Yrs)",
            "Location",
            "Evaluation Date",
        ]
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for s in scores:
            ws.append([
                s.candidate.full_name,
                s.candidate.email or "N/A",
                s.overall_score,
                s.mandatory_skills_score,
                s.experience_score,
                s.candidate.total_experience_years or 0.0,
                s.candidate.location or "N/A",
                s.created_at.strftime("%Y-%m-%d"),
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def generate_pdf_report(job: Job, score: Score) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("TitleStyle", parent=styles["Heading1"], fontSize=20, leading=24, textColor=colors.HexColor("#1E3A8A"))
        subtitle_style = ParagraphStyle("SubTitleStyle", parent=styles["Normal"], fontSize=12, textColor=colors.HexColor("#4B5563"))

        story.append(Paragraph(f"AI Candidate Assessment Report", title_style))
        story.append(Paragraph(f"Job Role: <b>{job.title}</b> | Candidate: <b>{score.candidate.full_name}</b>", subtitle_style))
        story.append(Spacer(1, 15))

        # Overall Score Box Table
        score_data = [
            ["Overall AI Score", f"{score.overall_score:.1f} / 100"],
            ["Mandatory Skills Match", f"{score.mandatory_skills_score:.1f}%"],
            ["Experience Match", f"{score.experience_score:.1f}%"],
            ["Total Experience", f"{score.candidate.total_experience_years or 0.0} Years"],
        ]
        score_table = Table(score_data, colWidths=[200, 200])
        score_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F3F4F6")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1F2937")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ])
        )
        story.append(score_table)
        story.append(Spacer(1, 15))

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def generate_comparison_pdf(job: Job, columns: list) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("TitleStyle", parent=styles["Heading1"], fontSize=18, leading=22, textColor=colors.HexColor("#1E3A8A"))
        subtitle_style = ParagraphStyle("SubTitleStyle", parent=styles["Normal"], fontSize=11, textColor=colors.HexColor("#4B5563"))
        header_style = ParagraphStyle("HeaderStyle", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold", textColor=colors.white)
        cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=9, leading=12, textColor=colors.HexColor("#1F2937"))
        bold_cell_style = ParagraphStyle("BoldCellStyle", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", leading=12, textColor=colors.HexColor("#1F2937"))

        story.append(Paragraph(f"Candidate Comparison Matrix Report", title_style))
        story.append(Paragraph(f"Job Role: <b>{job.title}</b> | Department: <b>{job.department}</b>", subtitle_style))
        story.append(Spacer(1, 15))

        # Table headers
        table_headers = [Paragraph("<b>Criteria</b>", header_style)]
        for col in columns:
            table_headers.append(Paragraph(f"<b>{col.full_name}</b>", header_style))

        table_data = [table_headers]

        # Rows mapping
        rows = [
            ("Overall Score", lambda c: f"{c.overall_score:.1f} / 100", True),
            ("Mandatory Skills", lambda c: f"{c.mandatory_skills_score:.0f}%", False),
            ("Experience Yrs", lambda c: f"{c.total_experience_years:.1f} Yrs", False),
            ("Confidence Level", lambda c: f"{c.confidence_score:.0f}%", False),
            ("Hiring Risks / Alerts", lambda c: "\n".join([f"- {r}" for r in c.risks]) if c.risks else "No major risk flags detected.", False),
        ]

        for label, extractor, is_bold in rows:
            row_data = [Paragraph(f"<b>{label}</b>", bold_cell_style if is_bold else cell_style)]
            for col in columns:
                val = extractor(col)
                style_to_use = bold_cell_style if is_bold else cell_style
                row_data.append(Paragraph(val.replace("\n", "<br/>"), style_to_use))
            table_data.append(row_data)

        col_width = 540 / (len(columns) + 1)
        widths = [col_width] * (len(columns) + 1)

        comp_table = Table(table_data, colWidths=widths)
        comp_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ])
        )

        for i in range(1, len(table_data)):
            bg_color = colors.HexColor("#F9FAFB") if i % 2 == 0 else colors.white
            comp_table.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), bg_color)]))

        story.append(comp_table)
        doc.build(story)
        return buffer.getvalue()
